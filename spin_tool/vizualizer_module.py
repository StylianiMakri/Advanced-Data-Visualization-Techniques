#hi

import sys
import json
import csv
import re
import os
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QLabel,
    QLineEdit, QPushButton, QHBoxLayout, QHeaderView, QTextEdit, QMessageBox, QFileDialog, QListWidget, QSplitter
)
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


DATA_DIR = "data"

class StepDetailDialog(QMessageBox):
    def __init__(self, step, code):
        super().__init__()
        self.setWindowTitle(f"Details for Step {step.get('step', '')}")
        text = "\n".join(f"{k}: {v}" for k, v in step.items()) + f"\n\nCode:\n{code}"
        self.setText(text)
        self.setStandardButtons(QMessageBox.StandardButton.Ok)


class SpinVisualizer(QWidget):
    def __init__(self, parsed_data):
        super().__init__()
        self.setWindowTitle("SPIN Execution Timeline Visualizer")
        self.resize(1000, 700)

        self.data = parsed_data
        self.trail = self.data.get("trail", [])
        self.errors_raw = self.data.get("errors", [])
        self.load_pml_lines()

        self.errors = []
        for err in self.errors_raw:
            if isinstance(err, str):
                step = None
                match = re.search(r"step\s*(\d+)", err, re.IGNORECASE)
                if match:
                    step = int(match.group(1))
                self.errors.append({"message": err, "step": step})
            elif isinstance(err, dict):
                self.errors.append(err)
            else:
                self.errors.append({"message": str(err), "step": None})

        layout = QVBoxLayout()

        search_layout = QHBoxLayout()
        search_label = QLabel("Search:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Type to filter by any text in Step, Process, Line, or Action")
        self.search_input.textChanged.connect(self.filter_table)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)

        self.export_csv_btn = QPushButton("Export XLSX")
        self.export_csv_btn.clicked.connect(self.export_xlsx)
        self.export_html_btn = QPushButton("Export HTML")
        self.export_html_btn.clicked.connect(self.export_html)
        search_layout.addWidget(self.export_csv_btn)
        search_layout.addWidget(self.export_html_btn)

        layout.addLayout(search_layout)

        splitter = QSplitter(Qt.Orientation.Vertical)

        self.table = QTableWidget()
        self.table.setSortingEnabled(True)
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Step", "Process", "Line", "Action", "Code"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.cellDoubleClicked.connect(self.show_step_details)

        splitter.addWidget(self.table)

        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout()
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        bottom_layout.setSpacing(4)

        self.error_list = QListWidget()
        self.error_list.setMaximumHeight(100)

        error_label = QLabel("Errors Detected:")
        error_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        error_label.setStyleSheet("color: red; margin-bottom: 4px;")

        bottom_layout.addWidget(error_label)
        bottom_layout.addWidget(self.error_list)

        bottom_widget.setLayout(bottom_layout)
        splitter.addWidget(bottom_widget)

        layout.addWidget(splitter, stretch=1)
        self.setLayout(layout)

        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 1)

        self.load_data()
        self.populate_error_list()


    def load_data(self):
        self.table.setRowCount(len(self.trail))

        colors = [
            QColor("#FFC79A"), QColor("#83D9A1"), QColor("#677697"), QColor("#CE909E"),
            QColor("#626274"), QColor("#FFF278"), QColor("#4DCC5C"), QColor("#15B4AC"),
            QColor("#F56F6F"), QColor("#AD6ACA"), QColor("#175B1E"), QColor("#D47A1A"),
            QColor("#0B5D43"), QColor("#7EFFC7"), QColor("#D4CE1A"),
        ]

        for row, step in enumerate(self.trail):
            self.table.setItem(row, 0, QTableWidgetItem(str(step.get("step", ""))))
            proc_name = f'{step.get("proc_name", "")} (# {step.get("proc_id", "")})'
            proc_item = QTableWidgetItem(proc_name)
            color = colors[step.get("proc_id", 0) % len(colors)]
            proc_item.setBackground(color)
            self.table.setItem(row, 1, proc_item)

            self.table.setItem(row, 2, QTableWidgetItem(str(step.get("line", ""))))
            self.table.setItem(row, 3, QTableWidgetItem(step.get("action", "")))
            pml_line = self.get_pml_line(step.get("line", 0))
            self.table.setItem(row, 4, QTableWidgetItem(pml_line))


    def find_pml_file(self):
        for filename in os.listdir(DATA_DIR):
            if filename.endswith(".pml"):
                return os.path.join(DATA_DIR, filename)
        return None

    def load_pml_lines(self):
        pml_path = self.find_pml_file()
        if not pml_path:
            print("No .pml file found in data directory.")
            self.pml_lines = []
            return
        try:
            with open(pml_path, "r", encoding="utf-8") as f:
                self.pml_lines = f.readlines()
            print(f"Loaded {len(self.pml_lines)} lines from {pml_path}")
        except Exception as e:
            print(f"Failed to read .pml file: {e}")
            self.pml_lines = []

    def get_pml_line(self, line_number):
        if not isinstance(line_number, int) or line_number <= 0:
            return "— (invalid line)"
        if line_number > len(self.pml_lines):
            return "— (line not found in .pml)"

        raw = self.pml_lines[line_number - 1].strip()

        if raw.startswith('//') or raw.startswith('#') or raw == '':
            return "— (not executable code)"

        return raw



    def filter_table(self):
        filter_text = self.search_input.text().lower()
        for row in range(self.table.rowCount()):
            match = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and filter_text in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(row, not match)

    def show_step_details(self, row, column):
        step_num_item = self.table.item(row, 0)
        if not step_num_item:
            return
        try:
            step_num = int(step_num_item.text())
        except ValueError:
            return
        step = next((s for s in self.trail if s.get("step") == step_num), None)
        if step:
            code = self.get_pml_line(step.get("line", 0))
            dlg = StepDetailDialog(step, code)
            dlg.exec()

    def export_xlsx(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save XLSX", "", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "SPIN Execution"

            headers = ["Step", "Process", "Line", "Action", "Code"]
            ws.append(headers)

            bold_font = Font(bold=True)
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="center")

            for step in self.trail:
                proc_name = f'{step.get("proc_name", "")} (# {step.get("proc_id", "")})'
                code = self.get_pml_line(step.get("line", 0)).replace("\n", " ").replace("\r", " ")
                row = [
                    step.get("step", ""),
                    proc_name,
                    step.get("line", ""),
                    step.get("action", ""),
                    code
                ]
                ws.append(row)

            for col_idx, _ in enumerate(headers, start=1):
                max_length = max(
                    len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
                    for row_idx in range(1, ws.max_row + 1)
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

            wb.save(path)
            QMessageBox.information(self, "Export XLSX", f"Excel file exported successfully to:\n{path}")

        except Exception as e:
            QMessageBox.warning(self, "Export XLSX", f"Failed to export XLSX:\n{e}")


    def export_html(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save HTML", "", "HTML Files (*.html *.htm)")
        if not path:
            return
        try:
            html = """
            <html><head><meta charset=\"UTF-8\"><style>
            table {border-collapse: collapse; width: 100%;}
            th, td {border: 1px solid #999; padding: 0.5em; text-align: left;}
            th {background-color: #eee;}
            </style></head><body>
            <h2>SPIN Execution Timeline</h2>
            <table>
            <thead><tr><th>Step</th><th>Process</th><th>Line</th><th>Action</th><th>Code</th></tr></thead>
            <tbody>
            """
            for step in self.trail:
                proc_name = f'{step.get("proc_name", "")} (# {step.get("proc_id", "")})'
                code = self.get_pml_line(step.get("line", 0)).replace("<", "&lt;").replace(">", "&gt;")
                html += f"<tr><td>{step.get('step','')}</td><td>{proc_name}</td><td>{step.get('line','')}</td><td>{step.get('action','')}</td><td>{code}</td></tr>"
            html += """
            </tbody></table>
            <h3>Errors Detected:</h3>
            <pre style=\"color:red;\">""" + "\n".join(err["message"].replace("<","&lt;").replace(">","&gt;") for err in self.errors) + "</pre>" + """
            </body></html>
            """
            with open(path, 'w', encoding='utf-8') as f:
                f.write(html)
            QMessageBox.information(self, "Export HTML", f"HTML exported successfully to:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Export HTML", f"Failed to export HTML:\n{e}")

    def populate_error_list(self):
        self.error_list.clear()
        for err in self.errors:
            step_info = f"Step {err['step']}" if err['step'] is not None else "No step info"
            item_text = f"{step_info}: {err['message'][:80]}{'...' if len(err['message']) > 80 else ''}"
            self.error_list.addItem(item_text)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    try:
        with open("output/parsed_data.json") as f:
            data = json.load(f)
    except Exception as e:
        data = {"trail": [], "errors": [f"Error loading data: {e}"]}

    window = SpinVisualizer(data)
    window.show()
    sys.exit(app.exec())
